import os
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

async def generate_match_xlsx(match_vs):
    """Generate Excel file for tennis match analysis."""
    date = datetime.now().strftime("%d/%m/%Y")
    tournament = match_vs[0]
    player1, player2 = match_vs[2], match_vs[3]

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    xlsx_data = _prepare_match_data(match_vs, player1, player2)
    _populate_sheet_with_data(sheet, xlsx_data)
    _adjust_column_widths(sheet)
    _apply_cell_borders(sheet)
    
    file_path = _save_workbook(workbook, date, tournament, player1, player2)
    print(f"Excel file '{file_path}' has been generated.")

def _prepare_match_data(match_vs, player1, player2):
    """Prepare the data structure for the Excel sheet."""
    return [
        ["", player1, player2],
        ["Classement mondial", match_vs[4], match_vs[5]],
        ["% Victoires carrière", match_vs[17], match_vs[28]],
        ["% Victoires surface", match_vs[19], match_vs[30]],
        ["", "", ""],
        ["Ratio V/D carrière", match_vs[6], match_vs[7]],
        ["Ratio V/D 1 an", match_vs[8], match_vs[9]], 
        ["Ratio V/D surface", match_vs[10], match_vs[11]],
        ["% Sets gagnés", match_vs[12], match_vs[13]],
        ["", "", ""],
        ["% Victoires 1 an", match_vs[16], match_vs[27]],
        ["% Victoires 1 an surface", match_vs[18], match_vs[29]],
        ["% Victoires 50 derniers matchs", match_vs[15], match_vs[26]],
        ["% Victoires 10 derniers matchs", match_vs[14], match_vs[25]],
        ["Nombre matchs joués dernier mois", match_vs[20], match_vs[31]],
        ["Durée cumulée des derniers matchs tournoi en cours", match_vs[21], match_vs[32]], 
        ["", "", ""],
        ["% 1er service", match_vs[22], match_vs[33]],
        ["% points gagnés service", match_vs[23], match_vs[34]],
        ["% balles de break sauvées", match_vs[24], match_vs[35]],
        ["", "", ""],
        ["Pourcentage victoire", _calculate_win_percentage(match_vs[36]), _calculate_win_percentage(match_vs[37])],
        ["", "", ""],
        ["Côte estimée", match_vs[36], match_vs[37]],
    ]

def _calculate_win_percentage(odds_value):
    """Calculate win percentage from odds."""
    return round(90 / float(odds_value), 2)

def _populate_sheet_with_data(sheet, xlsx_data):
    """Add data rows to the Excel sheet."""
    for row in xlsx_data:
        sheet.append(row)

def _adjust_column_widths(sheet):
    """Auto-adjust column widths based on content."""
    for column in sheet.columns:
        max_length = _get_max_column_length(column)
        column_letter = get_column_letter(column[0].column)
        adjusted_width = max_length + 2
        sheet.column_dimensions[column_letter].width = adjusted_width

def _get_max_column_length(column):
    """Get the maximum character length in a column."""
    max_length = 0
    for cell in column:
        try:
            cell_length = len(str(cell.value))
            if cell_length > max_length:
                max_length = cell_length
        except:
            pass
    return max_length

def _apply_cell_borders(sheet):
    """Apply thin borders to all cells in the sheet."""
    thin_border = _create_thin_border()
    
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

def _create_thin_border():
    """Create a thin border style for cells."""
    return Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )

def _save_workbook(workbook, date, tournament, player1, player2):
    """Save the workbook to the appropriate directory structure."""
    folder_path = _create_match_folder(tournament)
    filename = _generate_filename(date, player1, player2)
    full_path = os.path.join(folder_path, filename)
    
    workbook.save(full_path)
    return full_path

def _create_match_folder(tournament):
    """Create the directory structure for match files."""
    year = datetime.now().year
    month = datetime.now().strftime("%B")
    
    folder_path = os.path.join("Matchs", str(year), month, tournament.replace(' ', '_'))
    os.makedirs(folder_path, exist_ok=True)
    return folder_path

def _generate_filename(date, player1, player2):
    """Generate the Excel filename based on date and player names."""
    safe_date = date.replace('/', '-')
    safe_player1 = player1.replace(' ', '_')
    safe_player2 = player2.replace(' ', '_')
    return f"{safe_date}_{safe_player1}_vs_{safe_player2}.xlsx"